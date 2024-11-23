import os
import pandas as pd
import glob
from openpyxl import Workbook
import json

# Define the base directory
base_dir = 'repos_v2'

def parse_impact(impact_str):
    try:
        impact_list = json.loads(impact_str.replace("'", '"'))
        return [item['softwareQuality'] for item in impact_list]
    except:
        return []

# Walk through all subdirectories in repos_v2
for user_dir in os.listdir(base_dir):
    user_path = os.path.join(base_dir, user_dir)
    if not os.path.isdir(user_path):
        continue

    # Create a new workbook for each user in write_only mode
    excel_file = os.path.join(user_path, f'{user_dir}_aggregated_report.xlsx')
    workbook = Workbook(write_only=True)

    # Create sheets for issues and consolidated data
    issues_sheet = workbook.create_sheet(title='Issues')
    consolidated_sheet = workbook.create_sheet(title='Consolidated')

    issues_header_written = False
    consolidated_header_written = False

    all_issues_data = []
    all_consolidated_data = []

    for repo_dir in os.listdir(user_path):
        repo_path = os.path.join(user_path, repo_dir)
        if not os.path.isdir(repo_path):
            continue

        # Find all CSV files in the repository directory
        csv_files = glob.glob(os.path.join(repo_path, '*.csv'))
        
        for csv_file in csv_files:
            if csv_file.endswith('_issues.csv'):
                sheet = issues_sheet
                header_written = issues_header_written
                df = pd.read_csv(csv_file)
                df['Repository'] = repo_dir
                all_issues_data.append(df)
            elif csv_file.endswith('_consolidated.csv'):
                sheet = consolidated_sheet
                header_written = consolidated_header_written
                df = pd.read_csv(csv_file)
                df['Repository'] = repo_dir
                all_consolidated_data.append(df)
            else:
                continue
            
            # Read the CSV file
            print(f"Processing file: {csv_file}")
            
            # Write the header if not written yet
            if not header_written:
                sheet.append(['Repository'] + df.columns.tolist())
                if sheet == issues_sheet:
                    issues_header_written = True
                else:
                    consolidated_header_written = True
            
            # Write the data
            for row in df.itertuples(index=False, name=None):
                sheet.append([repo_dir] + list(row))

    # Remove the default sheet created by openpyxl
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    # Save the Excel file for the user
    workbook.save(excel_file)
    print(f"Aggregated report for {user_dir} saved to {excel_file}")

    # Create condensed report
    condensed_excel_file = os.path.join(user_path, f'{user_dir}_condensed_report.xlsx')
    condensed_workbook = Workbook(write_only=True)

    # Condense issues data by Severity
    if all_issues_data:
        issues_df = pd.concat(all_issues_data, ignore_index=True)
        severity_counts = issues_df.groupby(['File Extension', 'Type', 'Severity']).size().unstack(fill_value=0)
        
        # Ensure all severity levels are present
        for severity in ['BLOCKER', 'CRITICAL', 'INFO', 'MAJOR', 'MINOR']:
            if severity not in severity_counts.columns:
                severity_counts[severity] = 0
        
        severity_counts = severity_counts[['BLOCKER', 'CRITICAL', 'INFO', 'MAJOR', 'MINOR']]
        severity_counts['Total Issues'] = severity_counts.sum(axis=1)
        severity_counts = severity_counts.reset_index()

        condensed_issues_sheet = condensed_workbook.create_sheet(title='Condensed Issues')
        condensed_issues_sheet.append(['File Extension', 'Type', 'BLOCKER', 'CRITICAL', 'INFO', 'MAJOR', 'MINOR', 'Total Issues'])
        for row in severity_counts.itertuples(index=False):
            condensed_issues_sheet.append(row)

        # Condense issues data by Impact
        issues_df['Impact'] = issues_df['Impact'].apply(parse_impact)
        impact_df = issues_df.explode('Impact')
        impact_counts = impact_df.groupby(['File Extension', 'Type', 'Impact']).size().unstack(fill_value=0)
        
        # Ensure all impact types are present
        for impact in ['RELIABILITY', 'MAINTAINABILITY', 'SECURITY']:
            if impact not in impact_counts.columns:
                impact_counts[impact] = 0
        
        impact_counts = impact_counts[['RELIABILITY', 'MAINTAINABILITY', 'SECURITY']]
        impact_counts['Total Issues'] = impact_counts.sum(axis=1)
        impact_counts = impact_counts.reset_index()

        condensed_impact_sheet = condensed_workbook.create_sheet(title='Condensed Issues by Impact')
        condensed_impact_sheet.append(['File Extension', 'Type', 'RELIABILITY', 'MAINTAINABILITY', 'SECURITY', 'Total Issues'])
        for row in impact_counts.itertuples(index=False):
            condensed_impact_sheet.append(row)

    # Condense consolidated data
    if all_consolidated_data:
        consolidated_df = pd.concat(all_consolidated_data, ignore_index=True)
        loc_summary = consolidated_df.groupby('File Extension')['Total LOC'].sum().reset_index()

        condensed_consolidated_sheet = condensed_workbook.create_sheet(title='Condensed Consolidated')
        condensed_consolidated_sheet.append(['File Extension', 'Total LOC'])
        for row in loc_summary.itertuples(index=False):
            condensed_consolidated_sheet.append(row)

    # Remove the default sheet created by openpyxl
    if 'Sheet' in condensed_workbook.sheetnames:
        del condensed_workbook['Sheet']

    # Save the condensed Excel file for the user
    condensed_workbook.save(condensed_excel_file)
    print(f"Condensed report for {user_dir} saved to {condensed_excel_file}")

print("All user reports have been generated.")